# -*- coding: utf-8 -*-

# Various functions and packages imported
import os
import string
import itertools
from music21 import *
from glob import glob
from lxml import etree
from collections import Counter
from openpyxl import Workbook
import winsound

# Declaration of lists
# Each list is filled and used within various functions
_corpusNames = []
_corpusTrees = []
_corpusElements = []
_corpusLyrics = [] 
_corpusWordCount = []
_posWordCount = []
_negWordCount = []
_miscWordCount = []
_corpusChordCount = []
_majorChordCount = []
_minorChordCount = []
_otherChordCount = []

# These two lists are used as a point of reference when sorting text into 'positive', 'negative', or 'other' affects.
_comparePositive = ['luja', 'halle', 'halleluja', 'leben', 'kann', 'redemisti', 'gut', 'frˆhlich', 'loben', 'dankbar', 'trost', 'himmel', 'alleluja', 'adoramus', 'miserere', 'grosser', 'ewig', 'glauben', 'ganze', 'singen', 'grosse', 'helfen', 'ewigkeit', 'behütet', 'bräutigam', 'benedicimus', 'sanctam', 'passionem', 'hallehalle', 'selig', 'geboren', 'erlöst', 'herzens', 'schöpfer', 'freundlich', 'grossen', 'engel', 'grossem', 'wahrheit', 'geborn', 'jungfrau', 'freuet', 'heiland', 'hallehalleluja', 'vertraut', 'lieben', 'erhalten', 'allerhöchsten', 'frommen', 'heiliger', 'freulande', 'herherzens', 'liebers', 'gegrüsset', 'lassen', 'gute', 'wonne', 'heilig', 'behüt', 'fromme', 'willen', 'lieber', 'ruhen', 'süsser', 'herzlich', 'lauter', 'mel', 'gehen', 'wohlthat', 'zuversicht', 'liewunben', 'hofalfet', 'gezieret', 'höchster', 'höchsten', 'neue', 'herrlichkeit', 'hoffnung', 'ergeben', 'lobet', 'heilgen', 'entgangen', 'alma', 'redemptoris', 'mater', 'succurre', 'ave', 'verdienet', 'himmelreich', 'höchmeistes', 'begnaladen', 'liebe', 'gnädiglich', 'gnädig', 'wohlgefällt', 'gerecht', 'herrlich', 'lebens', 'wepflege', 'rechten', 'behüte', 'hermehren', 'nahermen', 'vergeben', 'errett', 'verjüng', 'süsse', 'lieblich', 'laden', 'helle', 'gnaden', 'göttlich', 'lichtes', 'herzliebster', 'gelingen', 'geleben', 'bet']
_compareNegative = ['nicht', 'tod', 'ach', 'hilf', 'todes', 'sünde', 'allein', 'blut', 'banden', 'sünder', 'peccatorum', 'ende', 'armen', 'leiden', 'gefangen', 'warum', 'eleison', 'nimmermehr', 'schaden', 'vergiss', 'verlassen', 'bittern', 'derbar', 'vermeilangt', 'sterben', 'teufels', 'beschweret', 'niemand', 'wunden', 'eleis', 'altraulerigzeit', 'gewunbunden', 'schimpfiret', 'elend', 'lein', 'brunnschmerquell', 'armes', 'schweren', 'sünde', 'herhimze', 'winden', 'todesnoth', 'schwachheit', 'schanden', 'verbrochen', 'urtheil', 'missethaten', 'feinde']


# This declares the corpus for Music21
# The directory reference in _projectCorpus.addPath will need to be changed to where you saved the file
_projectCorpus = corpus.corpora.LocalCorpus()
_projectCorpus.addPath('E:/Documents/UNE/MUSI366/Python Projects/MUSI366-Project-Folder/Corpus')

# Defines spreadsheet which the data will be written to
workbook = Workbook()
sheet = workbook.active

# Defines 'progress' as a global variable
progress = 0

# This fills _corpusNames[] and _corpusTrees[] 
def ProcessCorpus(): 
	for filename in glob(r'E:/Documents/UNE/MUSI366/Python Projects/MUSI366-Project-Folder/Corpus\*.xml'):
		tree = etree.parse(filename)   
		_corpusTrees.append(tree)
		_corpusNames.append(os.path.basename(filename))
		
# This processes, collects, and sorts lyrics from the corpus XML files
def ProcessText():
	print("Processing text now.")   

	for filename in glob(r'E:/Documents/UNE/MUSI366/Python Projects/MUSI366-Project-Folder/Corpus\*.xml'):
		tree = etree.parse(filename) 
		root = tree.getroot()

		_textList = []
		_syllabicList = []
		_wordParts = []
		_wordList = []
		_posWords =[]
		_negWords = []
		_miscWords = []
		_wordCount = []      
			
		for text in root.iter('text'):
			_textList.append(text.text)
									
		for syllabic in root.iter('syllabic'):
			_syllabicList.append(syllabic.text)
		
		# This loop locates syllables and combines them into whole words
		# Completed words are aded to _wordList[]
		for text, syllabic in zip(_textList, _syllabicList):
			if(syllabic == 'begin'):
				_wordParts.append(text)
			elif(syllabic == 'middle'): 
				_wordParts.append(text)
			elif(syllabic == 'end'):
				_wordParts.append(text)
				_word = ''.join(_wordParts)
				_wordList.append(_word)
				_wordParts.clear()
			elif(syllabic == 'single'):
				_wordList.append(text)
				_wordParts.clear()
				
		# This converts all words to lowercase and removes punctuation
		_wordList = [''.join(character for character in word if character not in string.punctuation) for word in _wordList]
		_wordList = map(lambda x:x.lower(), _wordList)
		
		# This adds _wordList[] to _corpusLyrics[]
		_corpusLyrics.append(_wordList)
		
		# This sorts words by affect
		for word in _wordList:
			if word in _comparePositive:
				_posWords.append(word)
			elif word in _compareNegative:
				_negWords.append(word)
			else:
				_miscWords.append(word)
		
		_posWordCount.append(len(_posWords))
		_negWordCount.append(len(_negWords))
		_miscWordCount.append(len(_miscWords))
		_posWords.clear()
		_negWords.clear()
		_miscWords.clear()
		_wordCount.clear()

		ProgressReport()

	# This notifies the user when all the text has been processed    
	print("Text processed!")
	ProgressReset()

# This counts and prints total words of Corpus (Not currently in use)
# This function was used in to extract words for trasnlation
# Once translated, words were sorted by affect and added to _comparePositive[] and _compareNegative[]
def TextCount():
	_count = (list(itertools.chain.from_iterable(_corpusLyrics)))
	print('')
	print('Corpus total word count is: ', len(_count))
	print(Counter(_count))

# This 'chordifies' each song in the corpus, then sorts each chord by tonality    
def ProcessChords():
	print("Processing chords now. This may take some time...")
	
	for song in _corpusNames:
		_targetSong = corpus.parse(song)
		_targetChords = _targetSong.chordify()
		_majorChords = []
		_minorChords = []
		_otherChords = []
		_chordCount = []
		
		for chord in _targetChords.recurse().getElementsByClass('Chord'):          
			if chord.isMajorTriad():
				_majorChords.append(chord)
			elif chord.isMinorTriad():
				_minorChords.append(chord)
			else:
				_otherChords.append(chord)
				
		_majorChordCount.append(len(_majorChords))
		_minorChordCount.append(len(_minorChords))
		_otherChordCount.append(len(_otherChords))
		_majorChords.clear()
		_minorChords.clear()
		_otherChords.clear()
		_chordCount.clear()

		ProgressReport()

	# This notifies the user when the chords have all been processed    
	print("Chords processed!")
	ProgressReset()

# This fills in the exceln spreadsheet, ready to be used for analysis
def FillSpreadsheet():
	print("Creating spreadshet now.")
	workbook.create_sheet("Test Sheet")
	
	# Fills 'Title' column with song titles
	def FillTitles():
		cellTitle = sheet.cell(row = (data + 1), column = 1)
		cellTitle.value = _corpusNames[data]

	# Fills 'Major' column with total number of major chords in each song  
	def FillMajor():
		cellMajor = sheet.cell(row = (data + 1), column = 2)
		cellMajor.value = _majorChordCount[data]

	# Fills 'Minor' column with total number of minor chords in each song     
	def FillMinor():
		cellMinor = sheet.cell(row = (data + 1), column = 3)
		cellMinor.value = _minorChordCount[data]

	# Fills 'Other' column with total number of other chords in each song     
	def FillOther():
		cellOther = sheet.cell(row = (data + 1), column = 4)
		cellOther.value = _otherChordCount[data]

	# Fills 'Tonality' column after defining overall tonality of each song   
	def FillTonality():
		cellTonality = sheet.cell(row = (data + 1), column = 5)
		if _majorChordCount[data] > _minorChordCount[data]:
			cellTonality.value = "Major"
		elif _majorChordCount[data] < _minorChordCount[data]:
			cellTonality.value = "Minor"
		else:
			cellTonality.value = "Undetermined"

	# Fills 'Positive' column with total number of positive words in each song         
	def FillPositive():
		cellPositive = sheet.cell(row = (data + 1), column = 6)
		cellPositive.value = _posWordCount[data]

	# Fills 'Negative' column with total number of negative words in each song      
	def FillNegative():
		cellNegative = sheet.cell(row = (data + 1), column = 7)
		cellNegative.value = _negWordCount[data]

	# Fills 'Misc' column with total number of misc words in each song      
	def FillMisc():
		cellMisc = sheet.cell(row = (data + 1), column = 8)
		cellMisc.value = _miscWordCount[data]

	# Fills 'Affect' column after defining overall affect of each song      
	def FillAffect():
		cellAffect = sheet.cell(row = (data + 1), column = 9)
		if _posWordCount[data] > _negWordCount[data]:
			cellAffect.value = "Positive"
		elif _posWordCount[data] < _negWordCount[data]:
			cellAffect.value = "Negative"
		else:
			cellAffect.value = "Undetermined"

	# Fills 'Congruency' column after defining whether the chords and words are congruent          
	def FillCongruency():
		cellCongruency = sheet.cell(row = (data + 1), column = 10)
		compCellTone = sheet.cell(row = (data + 1), column = 5)
		compCellWord = sheet.cell(row = (data + 1), column = 9)
		if compCellTone.value == "Major" and compCellWord.value == "Positive":
			cellCongruency.value = "Congruent"
		elif compCellTone.value == "Minor" and compCellWord.value == "Negative":
			cellCongruency.value = "Congruent"
		elif compCellTone.value == "Major" and compCellWord.value == "Negative":
			cellCongruency.value = "Incongruent"
		elif compCellTone.value == "Minor" and compCellWord.value == "Positive":
			cellCongruency.value = "Incongruent"
		else:
			cellCongruency.value = "Undetermined"

	# Inserts a new row and fills it with relevent headings of each column        
	def FillHeadings():
		sheet.insert_rows(1)
		sheet.cell(row = 1, column = 1).value = "Title"
		sheet.cell(row = 1, column = 2).value = "Major"
		sheet.cell(row = 1, column = 3).value = "Minor"
		sheet.cell(row = 1, column = 4).value = "Other"
		sheet.cell(row = 1, column = 5).value = "Tonality"
		sheet.cell(row = 1, column = 6).value = "Positive"
		sheet.cell(row = 1, column = 7).value = "Negative"
		sheet.cell(row = 1, column = 8).value = "Misc"
		sheet.cell(row = 1, column = 9).value = "Affect"
		sheet.cell(row = 1, column = 10).value = "Congruency"
   
	# Initiates all nested functions
	for data in range(0, 449):
		FillTitles()
		FillMajor()
		FillMinor()
		FillOther()
		FillTonality()
		FillPositive()
		FillNegative()
		FillMisc()
		FillAffect()
		FillCongruency()
		
	FillHeadings()
	# This value should be incremented in order to keep a record of previous outputs   
	workbook.save('Output_02.xlsx')

   # This notifies the user that the spreadsheet has been completed 
	print("Spreadsheet complete!")

# This reports the progress of some functions to the user
def ProgressReport():
	global progress
	progress += 1
	print("Files Processed: ", progress, "/449")

# This resets 'progress' to 0
def ProgressReset():
	global progress
	progress= 0

# This function alerts the user when the program has completed
def SignalCompletion():
	frequency = 1200  # Set Frequency To 1200 Hertz
	duration = 250  # Set Duration To 250 ms == 1/4 second

	print("The document is now ready for further analysis!")
	winsound.Beep(frequency, duration)

# Initiates all functions in sequence
ProcessCorpus()
ProcessText()
ProcessChords()
FillSpreadsheet()
SignalCompletion()